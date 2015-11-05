#!/user/bin/python3
# -*- coding: utf-8 -*-
# compatible with Python 3.4.3

__author__="Bo Zhou"
__copyright__ = "Copyright 2015, Dalewareriverkeeper project "
__credits__ = ["Bo Zhou"]
__license__ = "MIT"
__version__ = "1.0.0"
__maintainer__ = "Bo Zhou"
__email__ = "bzhou2@ualberta.ca"
__status__ = "Testing"

import time
import bs4
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter 
from openpyxl import cell
import geopy
import requests
    
def show_time(time):
    hours = time//3600
    minutes = (time//60)%60
    seconds = time%60
    print ("program runs for "+str(int(hours))+" hours, "+str(int(minutes))+" minutes, "+str(seconds)+" seconds.")

def write_to_excel(allData):
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    ws = wb.create_sheet()
    ws.title = "Event Data"
    row = 1
    for data in allData:
        column = 1
        for el in data:
            ws.cell(row,column).value = el
            column += 1            
        row += 1
    ew.save(filename=r"test.xlsx")
    return


# get whole page data
# return: [startDate, EndDate, Url, name, time, location, description]
def catch_data(response):
    r = response.url
    #print(r)
    soup = bs4.BeautifulSoup(response.content)
    pageContent = soup.find(id = "pageContent")
    allInfo = pageContent.find_all("tr")
    #print(pageContent)
    result = []
    # deal with date:
    dateList = allInfo[1].text.split()
    result.append(dateList[1])
    try:
        result.append(dateList[3])
    except IndexError:
        result.append(dateList[1])
    # add event url
    result.append(r)
    # add event name
    result.append(allInfo[0].text)
    # add event time
    timeList = allInfo[2].text.split()
    timeList.pop(0)
    time = " ".join(timeList)
    result.append(time)
    # add location:
    locationList = allInfo[3].text.split()
    #print(locationList)
    for i in range(2):
        locationList.pop(0)
    location = " ".join(locationList)
    result.append(location)
    # add description:
    result.append(allInfo[4].text)
    # print(dateList)
    return result


if __name__ == '__main__':
    startTime = time.time()
    saveFileName = "Daleware"
    crawlUrl = 'http://www.delawareriverkeeper.org/about/event.aspx?Id='
    pageId = 0
    allData = []
    nodata = 0
    while True:
        pageUrl = crawlUrl + str(pageId)
        try:
            response = requests.get(pageUrl)
        except Exception as e:
            # e is 'Connection aborted.'
            print("cannot open")
        if response.status_code == 200:
            pageData = catch_data(response)
            #print(pageData)
            allData.append(pageData)
            print ("opening page: "+str(pageId))
            nodata = 0
        elif response.status_code == 500:
            print ("No result on page: "+str(pageId))
            nodata += 1
        else:
            print ("Cannot open page: "+str(pageId))
        pageId += 1
        if nodata == 20:
            break
    write_to_excel(allData)
    #print(response.status_code)
    elapsedTime = time.time() - startTime
    show_time(elapsedTime)   
    print('all finish! ')